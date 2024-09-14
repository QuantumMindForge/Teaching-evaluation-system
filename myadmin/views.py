import hashlib
import os
from django.db.models import Avg
import openpyxl
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, HttpResponse, get_object_or_404

from login.models import *
from django import template

register = template.Library()

@register.filter
def mod(value, arg):
    """Returns value % arg"""
    return value % arg


# Create your views here.

# 管理员首页
def myadmin(request):
    # 这些代码是一个 Django ORM（对象关系映射）查询，具体用于查询并计算一个满足特定条件的记录的数量
    # 这些代码的作用是查询 这些 表中所有 is_active 字段为 True 的记录的数量，并将这个数量赋值给一个变量。
    # 简而言之，它是在统计活跃状态的学生数量，也正是因为这里，is_active才起到了一个逻辑删除的作用。
    students = Students.objects.filter(is_active=True).count()
    teachers = Teachers.objects.filter(is_active=True).count()
    kecheng = KeCheng.objects.filter(is_active=True).count()
    tiku = TiKu.objects.filter(is_active=True).count()
    pingjia = PingJia.objects.filter(is_active=True).count()
    admin = GuanLiYuan.objects.filter(is_active=True).count()

    return render(request, 'myadmin/index.html', context={'students': students,
                                                          'teachers': teachers,
                                                          'kecheng': kecheng,
                                                          'tiku': tiku,
                                                          'pingjia': pingjia,
                                                          'admin': admin})


def pswd_update(request):
    if request.method == 'GET':
        return render(request, 'myadmin/pswd_update.html')
    if request.method == 'POST':
        name = request.session.get('name')
        pswd = request.POST['pswd']
        pswd_1 = request.POST['pswd_1']
        pswd_2 = request.POST['pswd_2']

        if pswd_1 != pswd_2:
            msg = "新密码不一致！！！"
            return render(request, 'myadmin/pswd_update.html', {"msg": msg})
        try:
            ss = GuanLiYuan.objects.filter(name=name, password=pswd, is_active=True)
            sm = GuanLiYuan.objects.get(name=name, is_active=True)
        except Exception as e:
            print("myadmin_update_pswd:", e)
        if ss:
            sm.password = pswd_1
            sm.save()
            msg = "密码修改成功！！"
            return render(request, 'myadmin/index.html', {"msg": msg})
        else:
            msg = "原密码错误！"
            return render(request, 'myadmin/pswd_update.html', {"msg": msg})


# 管理学生界面
# pIndex是分页逻辑，表示当前页数，便于数据的分页展示
def myadmin_stu(request, pIndex=1):
    stu_list = Students.objects.filter().order_by()
    mywhere = []

    # 获取查询参数
    kw = request.GET.get("keyword", None)
    if kw:
        # 条件获取，即查询学号或姓名符合查询参数kw的值,并将其放到mywhere列表中
        stu_list = stu_list.filter(Q(student_id__contains=kw) | Q(name__contains=kw))
        mywhere.append('keyword' + kw)

    # 对数据进行分页处理
    pIndex = int(pIndex)
    # 按照每一页9条数据的形式进行显示
    page = Paginator(stu_list, 10)
    max_page = page.num_pages

    # 防止越界
    if pIndex > max_page:
        pIndex = max_page
    if pIndex < 1:
        pIndex = 1

    list1 = page.page(pIndex)
    plist = page.page_range

    return render(request, 'myadmin/students/myadmin_stu.html', context={'stulist': list1,
                                                         'plist': plist,
                                                         'pIndex': pIndex,
                                                         'max_page': max_page,
                                                         'mywhere': mywhere})

# 添加学生
def stu_add(request):
    if request.method == 'POST':
        student_id = request.POST['student_id']
        name = request.POST['name']
        phone = request.POST['phone']
        email = request.POST['email']
        xueyuan = request.POST['xueyuan']
        banji = request.POST['banji']
        sex = request.POST['sex']
        password = request.POST['password']

        try:
            stu = Students.objects.filter(student_id=student_id)
        except Exception as e:
            print(e)
        if stu:
            msg = "学号已存在，请检查您的学号是否正确！"
            return render(request, 'myadmin/students/stu_add.html', context={'msg': msg})
        else:
            stu_add = Students.objects.create(student_id=student_id, name=name, phone=phone, email=email, banji=banji, sex=sex, password=password, xueyuan=xueyuan)
            msg = '添加成功'
            return render(request, 'myadmin/tishi.html', context={'msg': msg})
    elif request.method == 'GET':
        return render(request, 'myadmin/students/stu_add.html')

# 编辑学生信息
def stu_edit(request, student_id):
    try:
        # 先try检查学生是否存在
        stu = Students.objects.get(student_id=str(student_id))
    except Exception as e:
        print(e)
        msg = '没有该学生，请检查后输入！'
        return render(request, 'myadmin/tishi.html', context={'msg': msg})

    if request.method == 'POST':
        name = request.POST['name']
        phone = request.POST['phone']
        email = request.POST['email']
        banji = request.POST['banji']
        sex = request.POST['sex']
        is_active = request.POST['is_active']

        stu.name = name
        stu.phone = phone
        stu.email = email
        stu.banji = banji
        stu.sex = sex
        stu.is_active = is_active
        stu.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})
    elif request.method == 'GET':
        return render(request, 'myadmin/students/stu_edit.html', locals())

# 删除学生信息
def stu_del(request, student_id):
    try:
        # 先try检查学生是否存在
        stu = Students.objects.get(student_id=str(student_id))
    except Exception as e:
        print(e)
        msg = '没有该学生，请检查后输入！'
        return render(request, 'myadmin/tishi2.html', context={'msg': msg})

    stu.is_active = False
    stu.save()
    msg = '删除成功'
    return render(request, 'myadmin/tishi2.html', context={'msg': msg})

# 上传学生信息
# 通过表格形式添加学生信息
def stu_upload(request):
    """
    这段代码的主要功能是从上传的 Excel 文件中读取学生信息，
    并检查每个学生的学号是否已经存在于数据库中。如果存在，则标记为“已存在”；
    如果不存在，则标记为“可以添加”
    """
    if request.method == 'POST':
        # 用于处理上传的 Excel 文件并导入学生信息。它通过读取用户上传的 .xlsx 文件，检查每一行的学生学号是否已存在于数据库中，然后根据结果生成反馈
        xlsx = request.FILES['xlsx']
        file_name, file_extension = os.path.splitext(str(xlsx))

        if file_extension == '.xlsx':
            wb = openpyxl.load_workbook(xlsx)
            worksheet = wb.active

            # 对数据进行单元格遍历
            excel_data = list()

            a = 0
            b = 0
            rows = worksheet.iter_rows()
            for row in rows:
                row_data = list()
                if 1 == row[1].row:
                    continue
                student_id = str(row[1].value)

                try:
                    stu = Students.objects.filter(student_id=student_id)
                except Exception as e:
                    print(e)

                if stu:
                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("no")
                    a += 1
                else:
                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("ok")
                    b += 1

                excel_data.append(row_data)
            global global_excel_data
            global_excel_data = excel_data
            msg = "可以添加%s个学生，%s个学号已存在！" % (b, a)
            return render(request, 'myadmin/students/stu_upload_ok.html', {"excel_data": excel_data, "msg": msg, "a": a})
        else:
            msg = "请你上传Excel文件错误！（格式必须XXXXX.xlsx）"
            return render(request, 'myadmin/tishi2.html', {"msg": msg})

    elif request.method == 'GET':
        return render(request, 'myadmin/students/stu_upload.html')

# 查看上传的的学生信息
def stu_toupload(request):
    stu_list = list()
    try:
        for row in global_excel_data:
            i = global_excel_data.index(row)
            # if i == 0:
            #     continue
            xuehao = row[1]
            name = row[2]
            xueyuan = row[3]
            banji = row[4]
            sex = row[5]
            email = row[7]
            phone = row[6]
            password = row[8]
            stu = Students.objects.filter(student_id=xuehao)
            if stu:  # 跳出数据库已存在的学生
                stu_list.append(xuehao)
                continue

            Students.objects.create(student_id=xuehao,
                                    name=name,
                                    password=password,
                                    xueyuan=xueyuan,
                                    banji=banji,
                                    sex=sex,
                                    email=email,
                                    phone=phone)
        msg = "成功上传学生表！！"
        return render(request, 'myadmin/tishi3.html', {"msg": msg})
    except Exception as e:
        print(e)
    return HttpResponse('请返回重新提交表格!！')

# 教师管理
def myadmin_teachers(request, pIndex):
    tea_list = Teachers.objects.filter().order_by()  # 班级过滤器
    mywhere = []
    # 获取并判断搜索
    kw = request.GET.get("keyword", None)
    if kw:
        tea_list = tea_list.filter(Q(teacher_id__contains=kw) | Q(name__contains=kw))
        mywhere.append('keyword' + kw)

    # 执行分页处理
    pIndex = int(pIndex)
    page = Paginator(tea_list, 10)  # 以每页9条数据分页
    maxpagex = page.num_pages  # 获取最大页数
    # 判断当前页是否越界
    if pIndex > maxpagex:
        pIndex = maxpagex
    if pIndex < 1:
        pIndex = 1

    list2 = page.page(pIndex)  # 获取当前页数据
    plist = page.page_range  # 获取页码表信息

    return render(request, 'myadmin/teachers/myadmin_teacher.html', context={"stulist": list2,
                                                                             "plist": plist,
                                                                             "pIndex": pIndex,
                                                                             "max_pages": maxpagex,
                                                                             'mywehere': mywhere})

# 添加老师
def teachers_add(request):
    if request.method == "GET":
        return render(request, 'myadmin/teachers/teacher_add.html')
    if request.method == "POST":
        teacher_id = request.POST['teacher_id']
        name = request.POST['name']
        password = request.POST['password']
        email = request.POST['email']
        phone = request.POST['phone']
        sex = request.POST['sex']

        try:
            stu = Teachers.objects.filter(teacher_id=teacher_id)
        except Exception as e:
            print(e)
        if stu:
            msg = '教工号已存在！'

            return render(request, 'myadmin/teachers/teacher_add.html', {"msg": msg})
        else:
            stu_add = Teachers.objects.create(teacher_id=teacher_id, name=name, phone=phone, email=email, sex=sex, password=password)
            msg = '添加教师成功！'
            return render(request, 'myadmin/tishi.html', {"msg": msg})

def teachers_edit(request, teacher_id):
    try:
        tea = Teachers.objects.get(teacher_id=teacher_id)
    except Exception as e:
        print(e)
        msg = '没有此教师！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/teachers/teacher_edit.html', locals())
    else:

        name = request.POST['name']
        phone = request.POST['phone']
        email = request.POST['email']
        sex = request.POST['sex']
        is_active = request.POST['is_active']

        tea.name = name
        tea.phone = phone
        tea.email = email
        tea.sex = sex
        tea.is_active = is_active
        tea.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})

def teachers_del(request, teacher_id):
    try:
        tea = Teachers.objects.get(teacher_id=teacher_id)
    except Exception as e:
        print(e)
        msg = "没有此教师！"
        return render(request, 'myadmin/tishi2.html', {"msg": msg})
    tea.is_active = False
    tea.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi2.html', {"msg": msg})

def teachers_upload(request):
    if request.method == 'GET':
        return render(request, 'myadmin/teachers/teacher_upload.html')
    elif request.method == 'POST':
        xlsx = request.FILES['xlsx']

        file_name, file_extension = os.path.splitext(str(xlsx))
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(xlsx)

            worksheet = wb.active  # 当前活跃的表单
            # print(worksheet)

            # 单元格遍历
            excel_data = list()

            a = 0
            b = 0
            rows = worksheet.iter_rows()
            for row in rows:
                row_data = list()
                if 1 == row[1].row:
                    continue
                teacher_id = str(row[1].value)

                try:
                    stu = Teachers.objects.filter(teacher_id=teacher_id)
                except Exception as e:
                    print(e)
                if stu:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("no")
                    a += 1
                else:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("ok")
                    b += 1
                excel_data.append(row_data)
            global global_excel_teachers
            global_excel_teachers = excel_data
            msg = "可以添加%s个教师ID，%s个教师ID已存在！" % (b, a)
            return render(request, 'myadmin/teachers/teacher_upload_ok.html',
                          {"excel_data": excel_data, "msg": msg, "a": a})
        else:
            msg = "请你上传Excel文件错误！（格式必须XXXXX.xlsx）"
            return render(request, 'myadmin/tishi2.html', {"msg": msg})

def teachers_toupload(request):
    stu_list = list()
    try:
        for row in global_excel_teachers:
            i = global_excel_teachers.index(row)

            teacher_id = row[1]
            name = row[2]

            sex = row[3]
            email = row[4]
            phone = row[5]
            password = row[6]

            stu = Teachers.objects.filter(teacher_id=teacher_id)
            if stu:  # 跳出数据库已存在的学生
                stu_list.append(teacher_id)
                continue

            Teachers.objects.create(teacher_id=teacher_id,
                                    name=name,
                                    sex=sex,
                                    email=email,
                                    phone=phone,
                                    password=password)

        msg = "成功上传教师表！！"
        return render(request, 'myadmin/tishi3.html', {"msg": msg})
    except Exception as e:
        print("teacher_toupload:", e)
    return HttpResponse('请返回重新提交表格!！')

# 课程管理首页
def myadmin_kecheng(request, pIndex):
    stu_list = KeCheng.objects.filter().order_by().values('id', 'kecheng', 'student_id', 'student_id__name', 'teacher_id',

                                                          'teacher_id__name', 'is_active')  # 班级过滤器
    mywhere = []
    # 获取并判断搜索
    kw = request.GET.get("keyword", None)
    if kw:
        stu_list = stu_list.filter(Q(id__contains=kw) | Q(kecheng__contains=kw))
        mywhere.append('keyword' + kw)

    # 执行分页处理
    pIndex = int(pIndex)
    page = Paginator(stu_list, 10)  # 以每页9条数据分页
    maxpagex = page.num_pages  # 获取最大页数
    # 判断当前页是否越界
    if pIndex > maxpagex:
        pIndex = maxpagex
    if pIndex < 1:
        pIndex = 1

    list2 = page.page(pIndex)  # 获取当前页数据
    plist = page.page_range  # 获取页码表信息
    context = {"stulist": list2, "plist": plist, "pIndex": pIndex, "max_pages": maxpagex, 'mywehere': mywhere}
    return render(request, 'myadmin/kecheng/myadmin_kecheng.html', context)

def myadmin_kecheng_add(request):
    if request.method == "GET":
        students = Students.objects.filter(is_active=True)
        teachers = Teachers.objects.filter(is_active=True)
        kecheng = KeCheng.objects.filter(is_active=True)

        try:
            id = kecheng.last().id
            return render(request, 'myadmin/kecheng/kecheng_add.html',
                          {"teachers": teachers, "students": students, "id": int(id) + 1})
        except Exception as e:
            print(e)
        id = 0
        return render(request, 'myadmin/kecheng/kecheng_add.html',
                      {"teachers": teachers, "students": students, "id": id})

    if request.method == "POST":
        id = request.POST['id']
        kecheng = request.POST['kecheng']
        xuehao = request.POST['student_id']
        teacher_id = request.POST['teacher_id']

        try:
            stu = KeCheng.objects.filter(str(id))
        except Exception as e:
            print(e)

        stu_add = KeCheng.objects.create(id=str(id), kecheng=kecheng, student_id=Students.objects.get(student_id=xuehao),
                                         teacher_id=Teachers.objects.get(teacher_id=teacher_id))
        msg = '添加课程成功！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

def myadmin_kecheng_edit(request, kecheng_id):
    try:
        stu = KeCheng.objects.get(id=kecheng_id)
        teachers = Teachers.objects.filter(is_active=True)
        students = Students.objects.filter(is_active=True)

    except Exception as e:
        print(e)
        msg = '没有此课程！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/kecheng/kecheng_edit.html', locals())
    else:
        kecheng = request.POST['kecheng']
        student_id = request.POST['student_id']
        teacher_id = request.POST['teacher_id']
        is_active = request.POST['is_active']

        stu.kecheng = kecheng
        stu.student_id = Students.objects.get(student_id=student_id)
        stu.teacher_id = Teachers.objects.get(teacher_id=teacher_id)
        stu.is_active = is_active
        stu.save()
        msg = '课程修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})

def myadmin_kecheng_del(request, kecheng_id):
    try:
        stu = KeCheng.objects.get(id=kecheng_id)
    except Exception as e:
        print(e)
        msg = "没有此课程！"
        return render(request, 'myadmin/tishi2.html', {"msg": msg})
    stu.is_active = False
    stu.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi2.html', {"msg": msg})

def kecheng_upload(request):
    if request.method == 'GET':
        return render(request, 'myadmin/kecheng/kecheng_upload.html')
    elif request.method == 'POST':
        xlsx = request.FILES['xlsx']

        file_name, file_extension = os.path.splitext(str(xlsx))
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(xlsx)

            worksheet = wb.active  # 当前活跃的表单
            # print(worksheet)

            # 单元格遍历
            excel_data = list()

            a = 0
            b = 0
            rows = worksheet.iter_rows()
            for row in rows:
                row_data = list()
                if 1 == row[1].row:
                    continue
                id = str(row[1].value)

                try:
                    stu = KeCheng.objects.filter(id=id)
                except Exception as e:
                    print(e)
                if stu:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("no")
                    a += 1
                else:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("ok")
                    b += 1
                # print(xuehao, row[1].row) #显示行数
                excel_data.append(row_data)
            global global_excel_kecheng
            global_excel_kecheng = excel_data
            msg = "可以添加%s个课程，%s个课程已存在！" % (b, a)
            return render(request, 'myadmin/kecheng/kecheng_upload_ok.html',
                          {"excel_data": excel_data, "msg": msg, "a": a})
        else:
            msg = "请你上传Excel文件错误！（格式必须XXXXX.xlsx）"
            return render(request, 'myadmin/tishi2.html', {"msg": msg})

def kecheng_toupload(request):
    stu_list = list()
    try:
        for row in global_excel_kecheng:
            i = global_excel_kecheng.index(row)

            id = row[1]
            name = row[2]

            student_id = row[3]

            teacher_id = row[4]
            # print(id, name, xuehao, teacher_id)

            stu = KeCheng.objects.filter(id=str(id))
            if stu:  # 跳出数据库已存在的课程
                stu_list.append(id)
                continue

            KeCheng.objects.create(id=id,
                                   kecheng=name,
                                   student_id=Students.objects.get(student_id=str(student_id)),
                                   teacher_id=Teachers.objects.get(teacher_id=str(teacher_id)))

        msg = "成功上传教师表！！"
        return render(request, 'myadmin/tishi3.html', {"msg": msg})
    except Exception as e:
        print(e)
    return HttpResponse('请返回重新提交表格!！')

# 展示题库
def myadmin_tiku(request):
    tiku_list = TiKu.objects.filter().order_by('id')  # 获取活跃题库并按ID排序
    context = {"tiku_list": tiku_list}
    return render(request, 'myadmin/tiku/myadmin_tiku.html', context)

# 添加题库
def myadmin_tiku_add(request):
    if request.method == "GET":
        options_scores = [('a', 'A'), ('b', 'B'), ('c', 'C'), ('d', 'D')]

        try:
            # 从 TiKu 表中获取下一个 ID
            last_tiku = TiKu.objects.last()
            next_id = (last_tiku.id + 1) if last_tiku else 1
        except Exception as e:
            print(e)
            next_id = 1

        return render(request, 'myadmin/tiku/tiku_add.html',
                      {"id": next_id, "options_scores": options_scores})

    if request.method == "POST":
        try:
            # 获取表单数据
            id = request.POST['id']
            timu = request.POST['timu']

            # 获取选项及分数
            option_a = request.POST['option_a']
            score_a = int(request.POST['score_a'])
            option_b = request.POST['option_b']
            score_b = int(request.POST['score_b'])
            option_c = request.POST['option_c']
            score_c = int(request.POST['score_c'])
            option_d = request.POST['option_d']
            score_d = int(request.POST['score_d'])

            # 创建新的 TiKu 实例
            new_tiku = TiKu.objects.create(
                id=id,
                timu=timu,
                option_a=option_a,
                score_a=score_a,
                option_b=option_b,
                score_b=score_b,
                option_c=option_c,
                score_c=score_c,
                option_d=option_d,
                score_d=score_d
            )
            msg = '添加题目成功！'
        except Exception as e:
            msg = f"添加题目失败：{str(e)}"

        return render(request, 'myadmin/tishi.html', {"msg": msg})

# 编辑题库
def myadmin_tiku_edit(request, id):
    try:
        tiku = TiKu.objects.get(id=int(id))
    except Exception as e:
        print(e)
        msg = '没有此评价题！！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    # Define option letters and corresponding labels
    options_scores = [('a', 'A'), ('b', 'B'), ('c', 'C'), ('d', 'D')]

    if request.method == "GET":
        # Prepare options and scores as individual variables
        option_a = tiku.option_a
        option_b = tiku.option_b
        option_c = tiku.option_c
        option_d = tiku.option_d
        score_a = tiku.score_a
        score_b = tiku.score_b
        score_c = tiku.score_c
        score_d = tiku.score_d

        return render(request, 'myadmin/tiku/tiku_edit.html', {
            "tiku": tiku,
            "options_scores": options_scores,
            "option_a": option_a,
            "option_b": option_b,
            "option_c": option_c,
            "option_d": option_d,
            "score_a": score_a,
            "score_b": score_b,
            "score_c": score_c,
            "score_d": score_d
        })

    if request.method == "POST":
        # Get the updated values from the form
        timu = request.POST.get('timu')
        option_a = request.POST.get('option_a')
        score_a = request.POST.get('score_a')
        option_b = request.POST.get('option_b')
        score_b = request.POST.get('score_b')
        option_c = request.POST.get('option_c')
        score_c = request.POST.get('score_c')
        option_d = request.POST.get('option_d')
        score_d = request.POST.get('score_d')
        is_active = request.POST.get('is_active')

        # Update the model fields
        tiku.timu = timu
        tiku.option_a = option_a
        tiku.score_a = score_a
        tiku.option_b = option_b
        tiku.score_b = score_b
        tiku.option_c = option_c
        tiku.score_c = score_c
        tiku.option_d = option_d
        tiku.score_d = score_d
        tiku.is_active = is_active

        tiku.save()
        msg = '修改成功！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

# 删除题库
def myadmin_tiku_del(request, id):
    try:
        tiku = TiKu.objects.get(id=int(id))

    except Exception as e:
        print(e)
        msg = '没有此评价题！！'
        return render(request, 'myadmin/tishi2.html', {"msg": msg})
    tiku.is_active = False
    tiku.save()
    msg = '删除成功！'
    return render(request, 'myadmin/tishi2.html', {"msg": msg})

# 展示教师的相关评价信息，包括评价率和平均分数。
def myadmin_pingjia(request):
    teachers_list = Teachers.objects.filter(is_active=True)  # 获取活跃的老师
    data = []

    # 获取评价对象
    pingjia = PingJia.objects.filter(is_active=True)  # 获取所有活跃的评价

    for teacher in teachers_list:
        my_list = []
        my_list.append(teacher.teacher_id)  # 老师ID
        my_list.append(teacher.name)  # 老师姓名

        # 老师性别
        sex = "男" if teacher.sex == "male" else "女" if teacher.sex == "female" else "未知"
        my_list.append(sex)

        my_list.append(teacher.phone)  # 老师电话

        # 学生评价数量
        pingjia_ok = pingjia.filter(kecheng__teacher_id=teacher.teacher_id).count()

        # 课程数量
        kecheng_sum = KeCheng.objects.filter(is_active=True, teacher_id=teacher.teacher_id).count()

        try:
            pjl = pingjia_ok / kecheng_sum if kecheng_sum > 0 else 0
            pjl = f"{pjl * 10:.2f}%"
            my_list.append(pjl)  # 添加评价率
        except Exception as e:
            print("myadmin_pingjia-error:", e)
            my_list.append('未授课')

        # 获取该教师所有课程中学生的平均得分
        avg = pingjia.filter(kecheng__teacher_id=teacher.teacher_id).aggregate(Avg('score'))
        avg_score = avg['score__avg']  # 获取平均得分

        try:
            if avg_score is not None:
                my_list.append(f"{avg_score:.2f}")  # 平均得分
            else:
                my_list.append('未评价')
        except Exception as e:
            print(e)
            my_list.append('未评价')

        data.append(my_list)

    return render(request, 'myadmin/pingjia/myadmin_pingjia.html', {"data_list": teachers_list, "data": data})
def myadmin_pingjia_show(request):
    teacher_id = request.GET.get('teacher_id')
    teacher = get_object_or_404(Teachers, teacher_id=teacher_id)

    # 获取该教师的所有课程
    kechengs = KeCheng.objects.filter(teacher_id=teacher_id, is_active=True)

    # 获取课程对应的所有评价
    pingjiabaio = PingJia.objects.filter(is_active=True, kecheng__teacher_id=teacher_id)

    # 计算评价率
    pingjia_sum = pingjiabaio.count() / 10  # 评价总数
    print(pingjia_sum)
    PJL = float('%.2f' % (pingjia_sum / kechengs.count() * 100))

    # 计算各题目得分的平均值
    avg_scores = pingjiabaio.aggregate(
        avg_score=Avg('score')
    )

    # 综合评价
    s_avg = avg_scores['avg_score'] if avg_scores['avg_score'] is not None else 0
    s_avg = float('%.2f' % s_avg)

    context = {
        'teacher': teacher,
        'kechengs': kechengs,
        'pingjiabaio': pingjiabaio,
        'PJL': PJL,
        's_avg': s_avg * 10
    }

    return render(request, 'myadmin/pingjia/pingjia_show.html', context)

# 用于展示班级的评价统计信息。
def myadmin_pingjia_pjl(request):
    data_list = []
    # distinct用于去重
    banji_data = Students.objects.filter(is_active=True).values('banji').distinct()

    for banji_record in banji_data:
        banji = banji_record['banji']
        student_count = Students.objects.filter(is_active=True, banji=banji).count()
        evaluated_count = KeCheng.objects.filter(is_ping='ok', student_id__banji=banji).values("student_id").distinct().count()

        no_evaluated_count = student_count - evaluated_count
        try:
            evaluation_rate = evaluated_count / student_count
            evaluation_rate_str = f"{evaluation_rate:.2%}"
        except ZeroDivisionError:
            evaluation_rate_str = "0%"

        data_list.append([banji, student_count, evaluated_count, no_evaluated_count, evaluation_rate_str])

    return render(request, 'myadmin/pingjia/myadmin_pingjia_pjl.html', {"data_list": data_list})

# 展示未评价的学生名单及其相关信息
def myadmin_pingjia_not(request):
    # 记录所有未评价的学生
    not_pingjia = KeCheng.objects.filter(is_active=True, is_ping="no").order_by("student_id").distinct().values(
        "student_id",
        "student_id__name",
        "student_id__banji",
        "student_id__phone",
        "student_id__email"
    )

    data_list = []
    for record in not_pingjia:
        student_info = [
            record['student_id'],
            record['student_id__name'],
            record['student_id__banji'],
            record['student_id__phone'],
            record['student_id__email'],
            KeCheng.objects.filter(student_id=record['student_id'], is_ping="no").count()
        ]
        data_list.append(student_info)

    return render(request, 'myadmin/pingjia/myadmin_pingjia_not.html', {"data_list": data_list})

# 管理员设置
def myadmin_admin(request):
    data_list = GuanLiYuan.objects.filter(is_active=True).values("name", "phone", "email")

    return render(request, 'myadmin/admin/myadmin_admin.html', {"data_list": data_list})


def myadmin_admin_add(request):
    if request.method == "GET":
        return render(request, 'myadmin/admin/admin_add.html')
    if request.method == "POST":
        name = request.POST['name']
        phone = request.POST['phone']
        email = request.POST['email']
        password = request.POST['password']

        try:
            stu = GuanLiYuan.objects.filter(name=name)
        except Exception as e:
            print(e)
        if stu:
            msg = '用户已存在！'
            return render(request, 'myadmin/admin/admin_add.html', {"msg": msg})
        else:
            stu_add = GuanLiYuan.objects.create(name=name,
                                                phone=phone,
                                                email=email,
                                                password=password)
            msg = '添加成功！'
            return render(request, 'myadmin/tishi.html', {"msg": msg})


def myadmin_admin_edit(request, name):
    try:
        stu = GuanLiYuan.objects.get(name=name)
    except Exception as e:
        print(e)
        msg = '没有此管理员！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/admin/admin_edit.html', locals())
    else:

        phone = request.POST['phone']
        email = request.POST['email']

        stu.name = name
        stu.phone = phone

        stu.email = email
        stu.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})


def myadmin_admin_del(request, name):
    try:
        stu = GuanLiYuan.objects.get(name=name)
    except Exception as e:
        print(e)
        msg = "没有此管理员！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    stu.is_active = False
    stu.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi.html', {"msg": msg})
